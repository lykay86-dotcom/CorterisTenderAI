"""Tests for business workflow Excel export."""

from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from app.reporting.workflow_excel import WorkflowExcelExporter
from app.repositories.business_metrics import (
    BusinessAuditAction,
    BusinessAuditEvent,
    BusinessRecordKind,
    BusinessStatus,
    BusinessWorkflowRecord,
)


NOW = datetime(2026, 7, 11, 18, 30)


def _record(
    record_id: str,
    *,
    kind: BusinessRecordKind,
    archived: bool = False,
) -> BusinessWorkflowRecord:
    return BusinessWorkflowRecord(
        id=record_id,
        kind=kind.value,
        tender_id=f"T-{record_id}",
        title=f"Запись {record_id}",
        status=(
            BusinessStatus.READY.value
            if kind == BusinessRecordKind.PROPOSAL
            else BusinessStatus.ACTIVE.value
            if kind == BusinessRecordKind.PROJECT
            else BusinessStatus.REVIEW.value
        ),
        total=1_000_000,
        profit=200_000,
        margin_percent=20,
        file_path="proposal.docx",
        due_date="2026-07-20",
        created_at="2026-07-10T10:00:00",
        updated_at="2026-07-11T12:00:00",
        archived_at=("2026-07-11T17:00:00" if archived else ""),
    )


def test_export_creates_summary_registry_and_history(tmp_path) -> None:
    records = [
        _record(
            "proposal",
            kind=BusinessRecordKind.PROPOSAL,
        ),
        _record(
            "project",
            kind=BusinessRecordKind.PROJECT,
            archived=True,
        ),
    ]
    events = [
        BusinessAuditEvent(
            id="event-1",
            record_id="proposal",
            action=BusinessAuditAction.CREATED.value,
            occurred_at="2026-07-11T10:00:00",
            new_value="Запись proposal",
        ),
        BusinessAuditEvent(
            id="event-2",
            record_id="proposal",
            action=BusinessAuditAction.STATUS_CHANGED.value,
            occurred_at="2026-07-11T11:00:00",
            field="status",
            old_value=BusinessStatus.DRAFT.value,
            new_value=BusinessStatus.READY.value,
        ),
    ]
    target = tmp_path / "workflow.xlsx"

    result = WorkflowExcelExporter().export(
        target,
        records=records,
        events=events,
        filter_description="Тип: Все типы",
        exported_at=NOW,
    )

    assert result.path == target
    assert result.record_count == 2
    assert result.event_count == 2
    assert target.exists()

    workbook = load_workbook(target)
    assert workbook.sheetnames == [
        "Сводка",
        "Реестр",
        "Журнал изменений",
    ]

    registry = workbook["Реестр"]
    assert registry.max_row == 3
    assert registry["B2"].value in {
        "Коммерческое предложение",
        "Проект",
    }
    assert "WorkflowRegistryTable" in registry.tables

    history = workbook["Журнал изменений"]
    assert history.max_row == 3
    assert "WorkflowAuditTable" in history.tables
    assert history["E2"].value in {
        "Создание",
        "Смена статуса",
    }


def test_registry_contains_money_dates_and_archive_flag(
    tmp_path,
) -> None:
    target = tmp_path / "workflow.xlsx"
    WorkflowExcelExporter().export(
        target,
        records=[
            _record(
                "archived",
                kind=BusinessRecordKind.PROPOSAL,
                archived=True,
            )
        ],
        exported_at=NOW,
    )

    workbook = load_workbook(target)
    registry = workbook["Реестр"]

    assert registry["F2"].value == 1_000_000
    assert "₽" in registry["F2"].number_format
    assert registry["M2"].value == "Да"
    assert registry["N2"].value == datetime(
        2026,
        7,
        11,
        17,
        0,
    )


def test_empty_export_remains_valid(tmp_path) -> None:
    target = tmp_path / "empty.xlsx"

    result = WorkflowExcelExporter().export(
        target,
        records=[],
        events=[],
        exported_at=NOW,
    )

    assert result.record_count == 0
    workbook = load_workbook(target)
    assert workbook["Реестр"].max_row == 1
    assert workbook["Журнал изменений"].max_row == 1
