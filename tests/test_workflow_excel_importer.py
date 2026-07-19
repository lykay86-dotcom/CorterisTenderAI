"""Tests for validated business workflow Excel import."""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook, load_workbook

from app.reporting.workflow_excel import WorkflowExcelExporter
from app.reporting.workflow_excel_import import WorkflowExcelImporter
from app.repositories.business_metrics import (
    BusinessMetricsRepository,
    BusinessRecordKind,
    BusinessStatus,
)


HEADERS = [
    "ID записи",
    "Тип",
    "Тендер",
    "Наименование",
    "Статус",
    "Сумма, ₽",
    "Прибыль, ₽",
    "Маржа, %",
    "Срок",
    "Документ",
    "Архив",
]


def _workbook(path, rows) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_preview_reads_valid_exported_registry(tmp_path) -> None:
    path = tmp_path / "registry.xlsx"
    _workbook(
        path,
        [
            [
                "",
                "Коммерческое предложение",
                "T-81",
                "КП на видеонаблюдение",
                "Готово к отправке",
                1_000_000,
                200_000,
                20,
                "20.07.2026",
                "proposal.docx",
                "Нет",
            ]
        ],
    )

    preview = WorkflowExcelImporter().preview(path)

    assert preview.can_import
    assert len(preview.valid_rows) == 1
    row = preview.valid_rows[0]
    assert row.kind == BusinessRecordKind.PROPOSAL
    assert row.status == BusinessStatus.READY
    assert row.total == Decimal("1000000")
    assert row.due_date == "2026-07-20"


def test_preview_rejects_invalid_type_status_and_numbers(
    tmp_path,
) -> None:
    path = tmp_path / "invalid.xlsx"
    _workbook(
        path,
        [
            [
                "",
                "Неизвестный тип",
                "",
                "",
                "Странный статус",
                "abc",
                -10,
                20,
                "31.31.2026",
                "",
                "Нет",
            ]
        ],
    )

    preview = WorkflowExcelImporter().preview(path)

    assert not preview.can_import
    assert len(preview.invalid_rows) == 1
    messages = " ".join(issue.message for issue in preview.invalid_rows[0].issues)
    assert "Неизвестный тип" in messages
    assert "Не указан тендер" in messages
    assert "Неверная дата" in messages


def test_apply_creates_updates_and_archives_records(tmp_path) -> None:
    repository = BusinessMetricsRepository(tmp_path / "workflow.json")
    existing = repository.save_record(
        kind=BusinessRecordKind.PROPOSAL,
        tender_id="T-1",
        title="Старое КП",
        status=BusinessStatus.DRAFT,
        total=100_000,
    )

    path = tmp_path / "import.xlsx"
    _workbook(
        path,
        [
            [
                existing.id,
                "Коммерческое предложение",
                "T-1",
                "Обновлённое КП",
                "Готово к отправке",
                500_000,
                100_000,
                20,
                "2026-07-25",
                "new.docx",
                "Нет",
            ],
            [
                "",
                "Проект",
                "T-2",
                "Монтаж СКУД",
                "В работе",
                2_000_000,
                400_000,
                20,
                "",
                "",
                "Да",
            ],
        ],
    )

    importer = WorkflowExcelImporter()
    preview = importer.preview(
        path,
        existing_records=repository.list_records(include_archived=True),
    )
    result = importer.apply(preview, repository)

    assert result.created == 1
    assert result.updated == 1
    assert result.archived == 1
    assert not result.failures

    records = repository.list_records(include_archived=True)
    updated = next(item for item in records if item.id == existing.id)
    project = next(item for item in records if item.kind == BusinessRecordKind.PROJECT.value)
    assert updated.title == "Обновлённое КП"
    assert updated.status == BusinessStatus.READY.value
    assert project.is_archived


def test_duplicate_kind_and_tender_rows_are_rejected(tmp_path) -> None:
    path = tmp_path / "duplicates.xlsx"
    duplicate = [
        "",
        "Смета",
        "T-3",
        "Смета",
        "Черновик",
        100_000,
        20_000,
        20,
        "",
        "",
        "Нет",
    ]
    _workbook(path, [duplicate, duplicate])

    preview = WorkflowExcelImporter().preview(path)

    assert len(preview.invalid_rows) == 2
    assert not preview.can_import


def test_export_import_round_trip_prefers_exact_financial_metadata(tmp_path) -> None:
    repository = BusinessMetricsRepository(tmp_path / "source.json")
    record = repository.save_record(
        kind=BusinessRecordKind.PROPOSAL,
        tender_id="T-EXACT",
        title="Exact",
        status=BusinessStatus.READY,
        total=Decimal("0.10"),
        profit=Decimal("0.01"),
    )
    path = tmp_path / "exact.xlsx"
    WorkflowExcelExporter().export(path, records=[record])

    preview = WorkflowExcelImporter().preview(path)

    assert preview.can_import
    assert preview.rows[0].total == Decimal("0.10")
    assert preview.rows[0].profit == Decimal("0.01")
    assert preview.rows[0].margin_percent == Decimal("10.00")


def test_import_rejects_numeric_cell_conflicting_with_exact_metadata(tmp_path) -> None:
    repository = BusinessMetricsRepository(tmp_path / "source.json")
    record = repository.save_record(
        kind=BusinessRecordKind.PROPOSAL,
        tender_id="T-TAMPER",
        title="Tamper",
        status=BusinessStatus.READY,
        total=Decimal("0.10"),
        profit=Decimal("0.01"),
    )
    path = tmp_path / "tampered.xlsx"
    WorkflowExcelExporter().export(path, records=[record])
    workbook = load_workbook(path)
    workbook["Реестр"]["F2"] = 0.11
    workbook.save(path)

    preview = WorkflowExcelImporter().preview(path)

    assert not preview.can_import
    assert any("FinancialExact" in issue.message for issue in preview.rows[0].issues)


def test_import_rejects_manual_margin_conflict(tmp_path) -> None:
    path = tmp_path / "margin-conflict.xlsx"
    _workbook(
        path,
        [
            [
                "",
                "Коммерческое предложение",
                "T-MARGIN",
                "Conflict",
                "Готово к отправке",
                100,
                10,
                25,
                "",
                "",
                "Нет",
            ]
        ],
    )

    preview = WorkflowExcelImporter().preview(path)

    assert not preview.can_import
    assert any("Маржа не совпадает" in issue.message for issue in preview.rows[0].issues)
