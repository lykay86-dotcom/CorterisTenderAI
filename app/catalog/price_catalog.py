from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook

@dataclass(slots=True)
class CatalogItem:
    category: str
    name: str
    unit: str
    base_cost: float
    market_min: float
    market_max: float
    item_type: str
    note: str

class PriceCatalog:
    def __init__(self, path: Path):
        self.path = path
        self.items: list[CatalogItem] = []

    def load(self) -> list[CatalogItem]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb['Прайс']
        self.items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            self.items.append(CatalogItem(
                category=str(row[0] or ''), name=str(row[1]), unit=str(row[2] or 'шт.'),
                base_cost=float(row[3] or 0), market_min=float(row[4] or 0),
                market_max=float(row[5] or 0), item_type=str(row[6] or ''), note=str(row[7] or '')
            ))
        wb.close()
        return self.items

    def search(self, query: str, limit: int = 20) -> list[CatalogItem]:
        q = query.strip().lower()
        if not self.items:
            self.load()
        if not q:
            return self.items[:limit]
        return [x for x in self.items if q in x.name.lower() or q in x.category.lower()][:limit]
