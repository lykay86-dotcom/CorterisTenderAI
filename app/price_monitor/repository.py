from __future__ import annotations
from pathlib import Path
import csv
import json
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from .models import PriceOffer


class PriceOfferRepository:
    def __init__(self, storage: Path):
        self.storage = storage
        self.offers: list[PriceOffer] = []
        self.load()

    def load(self) -> None:
        if not self.storage.exists():
            self.offers = []
            return
        raw = json.loads(self.storage.read_text(encoding="utf-8"))
        self.offers = [PriceOffer(**x) for x in raw]

    def save(self) -> None:
        self.storage.parent.mkdir(parents=True, exist_ok=True)
        self.storage.write_text(
            json.dumps([x.to_dict() for x in self.offers], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def upsert(self, offer: PriceOffer) -> None:
        for i, current in enumerate(self.offers):
            if current.key == offer.key:
                self.offers[i] = offer
                self.save()
                return
        self.offers.append(offer)
        self.save()

    @staticmethod
    def _num(row: dict, *keys: str, default: float = 0.0) -> float:
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                try:
                    return float(str(value).replace(" ", "").replace(",", "."))
                except ValueError:
                    pass
        return default

    @staticmethod
    def _bool(row: dict, *keys: str, default: bool = False) -> bool:
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                return str(value).strip().lower() in {"1", "true", "да", "yes", "y", "есть"}
        return default

    def _offer_from_row(self, raw: dict) -> PriceOffer | None:
        row = {str(k).strip().lower(): v for k, v in raw.items()}
        supplier = str(row.get("поставщик") or row.get("supplier") or "Не указан").strip()
        brand = str(row.get("бренд") or row.get("brand") or "").strip()
        model = str(row.get("модель") or row.get("model") or "").strip()
        if not model:
            return None
        chars = {}
        raw_chars = row.get("характеристики") or row.get("characteristics")
        if isinstance(raw_chars, str) and raw_chars.strip():
            try:
                chars = json.loads(raw_chars)
            except json.JSONDecodeError:
                for part in raw_chars.split(";"):
                    if "=" in part:
                        k, v = part.split("=", 1)
                        chars[k.strip()] = v.strip()
        return PriceOffer(
            supplier=supplier,
            brand=brand,
            model=model,
            unit_price=self._num(row, "цена", "unit_price", "закупочная цена", "price"),
            currency=str(row.get("валюта") or row.get("currency") or "RUB"),
            vat_included=self._bool(row, "ндс включен", "vat_included", default=True),
            vat_percent=self._num(row, "ставка ндс", "vat_percent", default=22),
            delivery_cost=self._num(row, "доставка", "delivery_cost"),
            discount_percent=self._num(row, "скидка", "discount_percent"),
            minimum_order_qty=self._num(row, "минимальная партия", "minimum_order_qty", default=1),
            stock=self._num(row, "остаток", "stock"),
            lead_time_days=int(self._num(row, "срок поставки", "lead_time_days")),
            warranty_months=int(self._num(row, "гарантия", "warranty_months")),
            official_supply=self._bool(row, "официальная поставка", "official_supply"),
            source=str(row.get("источник") or row.get("source") or "Импорт прайса"),
            source_url=str(row.get("ссылка") or row.get("source_url") or ""),
            article=str(row.get("артикул") or row.get("article") or ""),
            category=str(row.get("категория") or row.get("category") or ""),
            characteristics=chars,
            certificate_available=self._bool(row, "сертификат", "certificate_available"),
        )

    def import_file(self, path: Path) -> int:
        suffix = path.suffix.lower()
        rows: list[dict] = []
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
        elif suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(x or "").strip() for x in next(ws.iter_rows(values_only=True))]
            rows = [dict(zip(headers, row)) for row in ws.iter_rows(values_only=True)]
        elif suffix == ".json":
            raw = json.loads(path.read_text(encoding="utf-8"))
            rows = raw if isinstance(raw, list) else raw.get("offers", [])
        elif suffix == ".xml":
            root = ET.parse(path).getroot()
            rows = [
                {child.tag: child.text or "" for child in node} for node in root.findall(".//offer")
            ]
        else:
            raise ValueError("Поддерживаются CSV, XLSX, JSON и XML")
        count = 0
        for row in rows:
            offer = self._offer_from_row(row)
            if offer and offer.unit_price >= 0:
                self.upsert(offer)
                count += 1
        return count

    def history(self, brand: str, model: str) -> list[PriceOffer]:
        values = [
            x
            for x in self.offers
            if x.brand.lower() == brand.lower() and x.model.lower() == model.lower()
        ]
        return sorted(values, key=lambda x: x.checked_at, reverse=True)
