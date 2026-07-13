from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
import csv
import json
from openpyxl import load_workbook


@dataclass(slots=True)
class EquipmentItem:
    category: str
    brand: str
    model: str
    article: str = ""
    supplier: str = ""
    purchase_price: float = 0.0
    retail_price: float = 0.0
    stock: float = 0.0
    lead_time_days: int = 0
    warranty_months: int = 0
    country: str = ""
    passport_url: str = ""
    certificate_path: str = ""
    characteristics: dict[str, str] | None = None

    @property
    def key(self) -> str:
        return f"{self.brand}|{self.model}|{self.article}".lower()


class EquipmentCatalog:
    def __init__(self, storage: Path):
        self.storage = storage
        self.items: list[EquipmentItem] = []
        self.load()

    def load(self) -> None:
        if not self.storage.exists():
            self.items = []
            return
        raw = json.loads(self.storage.read_text(encoding="utf-8"))
        self.items = [EquipmentItem(**x) for x in raw]

    def save(self) -> None:
        self.storage.parent.mkdir(parents=True, exist_ok=True)
        self.storage.write_text(
            json.dumps([asdict(x) for x in self.items], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def upsert(self, item: EquipmentItem) -> None:
        for i, current in enumerate(self.items):
            if current.key == item.key:
                self.items[i] = item
                self.save()
                return
        self.items.append(item)
        self.save()

    def import_file(self, path: Path) -> int:
        rows: list[dict[str, object]] = []
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(x or "").strip() for x in next(ws.iter_rows(values_only=True))]
            rows = [dict(zip(headers, row)) for row in ws.iter_rows(values_only=True)]
        else:
            raise ValueError("Поддерживаются CSV и XLSX")
        count = 0
        for row in rows:
            normalized = {str(k).strip().lower(): v for k, v in row.items()}
            brand = str(normalized.get("бренд") or normalized.get("brand") or "").strip()
            model = str(normalized.get("модель") or normalized.get("model") or "").strip()
            if not brand or not model:
                continue

            def num(*keys: str) -> float:
                value = next(
                    (normalized.get(k) for k in keys if normalized.get(k) not in (None, "")), 0
                )
                try:
                    return float(str(value).replace(" ", "").replace(",", "."))
                except ValueError:
                    return 0.0

            self.upsert(
                EquipmentItem(
                    category=str(
                        normalized.get("категория") or normalized.get("category") or "Прочее"
                    ),
                    brand=brand,
                    model=model,
                    article=str(normalized.get("артикул") or normalized.get("article") or ""),
                    supplier=str(normalized.get("поставщик") or normalized.get("supplier") or ""),
                    purchase_price=num("закупочная цена", "purchase_price", "цена"),
                    retail_price=num("розничная цена", "retail_price"),
                    stock=num("остаток", "stock"),
                    lead_time_days=int(num("срок поставки", "lead_time_days")),
                    warranty_months=int(num("гарантия", "warranty_months")),
                    country=str(normalized.get("страна") or normalized.get("country") or ""),
                )
            )
            count += 1
        return count

    def search(self, query: str, limit: int = 100) -> list[EquipmentItem]:
        q = query.strip().lower()
        if not q:
            return self.items[:limit]
        return [
            x
            for x in self.items
            if q in f"{x.category} {x.brand} {x.model} {x.article} {x.supplier}".lower()
        ][:limit]
