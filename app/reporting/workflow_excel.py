"""Excel export for the business workflow registry and audit log."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.financial import (
    MARGIN_CONTRACT_VERSION,
    NUMERIC_CONTRACT_VERSION,
    canonical_money,
    canonical_percentage,
)
from app.repositories.business_metrics import (
    BusinessAuditAction,
    BusinessAuditEvent,
    BusinessRecordKind,
    BusinessStatus,
    BusinessWorkflowRecord,
)


KIND_LABELS = {
    BusinessRecordKind.PROPOSAL.value: "Коммерческое предложение",
    BusinessRecordKind.ESTIMATE.value: "Смета",
    BusinessRecordKind.PROJECT.value: "Проект",
}

STATUS_LABELS = {
    BusinessStatus.DRAFT.value: "Черновик",
    BusinessStatus.REVIEW.value: "На проверке",
    BusinessStatus.APPROVED.value: "Согласовано",
    BusinessStatus.READY.value: "Готово к отправке",
    BusinessStatus.SENT.value: "Отправлено",
    BusinessStatus.ACCEPTED.value: "Принято заказчиком",
    BusinessStatus.PLANNED.value: "Запланировано",
    BusinessStatus.ACTIVE.value: "В работе",
    BusinessStatus.INSTALLATION.value: "Монтаж",
    BusinessStatus.COMMISSIONING.value: "Пусконаладка",
    BusinessStatus.COMPLETED.value: "Завершено",
    BusinessStatus.CANCELLED.value: "Отменено",
    BusinessStatus.BLOCKED.value: "Заблокировано",
}

ACTION_LABELS = {
    BusinessAuditAction.CREATED.value: "Создание",
    BusinessAuditAction.UPDATED.value: "Изменение",
    BusinessAuditAction.STATUS_CHANGED.value: "Смена статуса",
    BusinessAuditAction.ARCHIVED.value: "Архивирование",
    BusinessAuditAction.RESTORED.value: "Восстановление",
}

FIELD_LABELS = {
    "title": "Название",
    "status": "Статус",
    "total": "Сумма",
    "profit": "Прибыль",
    "margin_percent": "Маржа",
    "file_path": "Файл",
    "due_date": "Срок",
    "archived_at": "Архив",
}


@dataclass(frozen=True, slots=True)
class WorkflowExcelExportResult:
    path: Path
    record_count: int
    event_count: int
    exported_at: datetime


class WorkflowExcelExporter:
    """Create a polished XLSX registry without modifying source templates."""

    DARK = "142B42"
    BLUE = "1479D1"
    LIGHT_BLUE = "EAF4FD"
    WHITE = "FFFFFF"
    TEXT = "17212B"
    MUTED = "667788"
    BORDER = "CFD8E3"
    GREEN = "DDF3E8"
    GREEN_TEXT = "147A4D"
    YELLOW = "FFF2CC"
    YELLOW_TEXT = "946200"
    RED = "FCE1E1"
    RED_TEXT = "A83232"
    GRAY = "E9EDF2"
    GRAY_TEXT = "65717E"

    REGISTRY_HEADERS = (
        "ID записи",
        "Тип",
        "Тендер",
        "Наименование",
        "Статус",
        "Сумма, ₽",
        "Прибыль, ₽",
        "Маржа, %",
        "Срок",
        "Документ",
        "Создано",
        "Обновлено",
        "Архив",
        "Дата архива",
    )

    HISTORY_HEADERS = (
        "Дата и время",
        "ID записи",
        "Тендер",
        "Наименование записи",
        "Действие",
        "Поле",
        "Старое значение",
        "Новое значение",
        "Пользователь",
    )

    def export(
        self,
        path: str | Path,
        *,
        records: Sequence[BusinessWorkflowRecord],
        events: Sequence[BusinessAuditEvent] = (),
        filter_description: str = "Все записи",
        exported_at: datetime | None = None,
    ) -> WorkflowExcelExportResult:
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)

        timestamp = exported_at or datetime.now()
        ordered_records = sorted(
            records,
            key=self._record_updated_at,
            reverse=True,
        )
        record_by_id = {record.id: record for record in ordered_records}
        ordered_events = sorted(
            (event for event in events if event.record_id in record_by_id),
            key=lambda event: event.timestamp,
            reverse=True,
        )

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Сводка"
        registry = workbook.create_sheet("Реестр")
        history = workbook.create_sheet("Журнал изменений")
        exact = workbook.create_sheet("FinancialExact")

        self._build_summary(
            summary,
            records=ordered_records,
            events=ordered_events,
            filter_description=filter_description,
            exported_at=timestamp,
        )
        self._build_registry(registry, ordered_records)
        self._build_history(
            history,
            ordered_events,
            record_by_id,
        )
        self._build_exact(exact, ordered_records)

        workbook.properties.title = "Реестр КП, смет и проектов — CORTERIS"
        workbook.properties.subject = "Выгрузка бизнес-процессов и журнала изменений"
        workbook.properties.creator = "Corteris Tender AI"
        workbook.properties.created = timestamp
        workbook.properties.modified = timestamp
        workbook.calculation.fullCalcOnLoad = True

        workbook.save(target)
        return WorkflowExcelExportResult(
            path=target,
            record_count=len(ordered_records),
            event_count=len(ordered_events),
            exported_at=timestamp,
        )

    @staticmethod
    def _build_exact(sheet, records: Sequence[BusinessWorkflowRecord]) -> None:
        """Embed authoritative fixed-point text beside Excel's numeric projection."""
        sheet.append(
            (
                "record_id",
                "total_exact",
                "profit_exact",
                "margin_exact",
                "currency",
                "money_unit",
                "margin_unit",
                "numeric_contract",
                "margin_contract",
            )
        )
        for record in records:
            sheet.append(
                (
                    record.id,
                    canonical_money(record.total),
                    canonical_money(record.profit),
                    canonical_percentage(record.margin_percent),
                    record.currency,
                    "money",
                    "percentage_point",
                    NUMERIC_CONTRACT_VERSION,
                    MARGIN_CONTRACT_VERSION,
                )
            )
        for row in sheet.iter_rows():
            for cell in row:
                cell.number_format = "@"
        sheet.sheet_state = "hidden"

    def _build_summary(
        self,
        sheet,
        *,
        records: Sequence[BusinessWorkflowRecord],
        events: Sequence[BusinessAuditEvent],
        filter_description: str,
        exported_at: datetime,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A6"

        sheet.merge_cells("A1:H2")
        title = sheet["A1"]
        title.value = "CORTERIS · Реестр КП, смет и проектов"
        title.font = Font(
            name="Arial",
            size=18,
            bold=True,
            color=self.WHITE,
        )
        title.fill = PatternFill("solid", fgColor=self.DARK)
        title.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        sheet["A3"] = "Дата выгрузки"
        sheet["B3"] = exported_at
        sheet["B3"].number_format = "dd.mm.yyyy hh:mm"
        sheet["D3"] = "Фильтр"
        sheet["E3"] = filter_description
        sheet.merge_cells("E3:H3")

        for cell in ("A3", "D3"):
            sheet[cell].font = Font(bold=True, color=self.MUTED)
        for cell in ("B3", "E3"):
            sheet[cell].font = Font(color=self.TEXT)

        active = [record for record in records if not record.is_archived]
        archived = [record for record in records if record.is_archived]
        proposals = [
            record for record in active if record.kind == BusinessRecordKind.PROPOSAL.value
        ]
        estimates = [
            record for record in active if record.kind == BusinessRecordKind.ESTIMATE.value
        ]
        projects = [record for record in active if record.kind == BusinessRecordKind.PROJECT.value]
        total_amount = sum(
            (Decimal(str(record.total)) for record in active),
            Decimal("0"),
        )
        total_profit = sum(
            (Decimal(str(record.profit)) for record in active),
            Decimal("0"),
        )

        kpis = (
            ("Всего записей", len(records), "0"),
            ("Активные", len(active), "0"),
            ("Архив", len(archived), "0"),
            ("События журнала", len(events), "0"),
            ("Общая сумма", float(total_amount), '#,##0.00 "₽"'),
            (
                "Потенциальная прибыль",
                float(total_profit),
                '#,##0.00 "₽"',
            ),
        )

        for index, (label, value, number_format) in enumerate(kpis):
            row = 6 + (index // 3) * 3
            column = 1 + (index % 3) * 3

            label_cell = sheet.cell(row=row, column=column)
            value_cell = sheet.cell(row=row + 1, column=column)
            sheet.merge_cells(
                start_row=row,
                start_column=column,
                end_row=row,
                end_column=column + 1,
            )
            sheet.merge_cells(
                start_row=row + 1,
                start_column=column,
                end_row=row + 1,
                end_column=column + 1,
            )

            label_cell.value = label
            label_cell.font = Font(
                bold=True,
                size=10,
                color=self.MUTED,
            )
            label_cell.fill = PatternFill(
                "solid",
                fgColor=self.LIGHT_BLUE,
            )
            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

            value_cell.value = value
            value_cell.number_format = number_format
            value_cell.font = Font(
                bold=True,
                size=16,
                color=self.TEXT,
            )
            value_cell.fill = PatternFill(
                "solid",
                fgColor=self.LIGHT_BLUE,
            )
            value_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

            for current_row in range(row, row + 2):
                for current_col in range(column, column + 2):
                    sheet.cell(
                        current_row,
                        current_col,
                    ).border = self._thin_border()

        count_start = 14
        sheet[f"A{count_start}"] = "Тип"
        sheet[f"B{count_start}"] = "Количество"
        type_counts = (
            ("Коммерческие предложения", len(proposals)),
            ("Сметы", len(estimates)),
            ("Проекты", len(projects)),
            ("Архивные записи", len(archived)),
        )
        for row_offset, values in enumerate(type_counts, 1):
            sheet.cell(count_start + row_offset, 1, values[0])
            sheet.cell(count_start + row_offset, 2, values[1])

        self._style_header(sheet[f"A{count_start}:B{count_start}"])
        for row in sheet.iter_rows(
            min_row=count_start + 1,
            max_row=count_start + len(type_counts),
            min_col=1,
            max_col=2,
        ):
            for cell in row:
                cell.border = self._thin_border()
                cell.alignment = Alignment(vertical="center")

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Распределение записей"
        chart.y_axis.title = "Количество"
        chart.x_axis.title = "Тип"
        chart.height = 7.2
        chart.width = 13.5
        data = Reference(
            sheet,
            min_col=2,
            min_row=count_start,
            max_row=count_start + len(type_counts),
        )
        categories = Reference(
            sheet,
            min_col=1,
            min_row=count_start + 1,
            max_row=count_start + len(type_counts),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        sheet.add_chart(chart, "D14")

        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 4
        for column in ("D", "E", "F", "G", "H"):
            sheet.column_dimensions[column].width = 17
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 28

    def _build_registry(
        self,
        sheet,
        records: Sequence[BusinessWorkflowRecord],
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.append(self.REGISTRY_HEADERS)
        self._style_header(sheet["1:1"])

        for record in records:
            row = [
                record.id,
                KIND_LABELS.get(record.kind, record.kind),
                record.tender_id,
                record.title,
                STATUS_LABELS.get(record.status, record.status),
                float(record.total),
                float(record.profit),
                float(record.margin_percent),
                self._parse_date(record.due_date),
                record.file_path,
                self._parse_datetime(record.created_at),
                self._parse_datetime(record.updated_at),
                "Да" if record.is_archived else "Нет",
                self._parse_datetime(record.archived_at),
            ]
            sheet.append(row)
            row_number = sheet.max_row

            for cell in sheet[row_number]:
                cell.border = self._thin_border()
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            sheet.cell(row_number, 6).number_format = '#,##0.00 "₽"'
            sheet.cell(row_number, 7).number_format = '#,##0.00 "₽"'
            sheet.cell(row_number, 8).number_format = '0.00"%"'
            sheet.cell(row_number, 9).number_format = "dd.mm.yyyy"
            sheet.cell(row_number, 11).number_format = "dd.mm.yyyy hh:mm"
            sheet.cell(row_number, 12).number_format = "dd.mm.yyyy hh:mm"
            sheet.cell(row_number, 14).number_format = "dd.mm.yyyy hh:mm"

            file_cell = sheet.cell(row_number, 10)
            if record.file_path:
                try:
                    file_cell.hyperlink = Path(record.file_path).expanduser().resolve().as_uri()
                    file_cell.style = "Hyperlink"
                except (OSError, ValueError):
                    pass

            if record.is_archived:
                for cell in sheet[row_number]:
                    cell.fill = PatternFill(
                        "solid",
                        fgColor=self.GRAY,
                    )
                    cell.font = Font(color=self.GRAY_TEXT)

        if records:
            table = Table(
                displayName="WorkflowRegistryTable",
                ref=f"A1:N{sheet.max_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

            sheet.conditional_formatting.add(
                f"E2:E{sheet.max_row}",
                FormulaRule(
                    formula=['OR(E2="Заблокировано",E2="Отменено")'],
                    fill=PatternFill("solid", fgColor=self.RED),
                    font=Font(color=self.RED_TEXT, bold=True),
                ),
            )
            sheet.conditional_formatting.add(
                f"E2:E{sheet.max_row}",
                FormulaRule(
                    formula=['OR(E2="На проверке",E2="Монтаж",E2="Пусконаладка")'],
                    fill=PatternFill("solid", fgColor=self.YELLOW),
                    font=Font(color=self.YELLOW_TEXT, bold=True),
                ),
            )
            sheet.conditional_formatting.add(
                f"E2:E{sheet.max_row}",
                FormulaRule(
                    formula=['OR(E2="Согласовано",E2="Завершено",E2="Принято заказчиком")'],
                    fill=PatternFill("solid", fgColor=self.GREEN),
                    font=Font(color=self.GREEN_TEXT, bold=True),
                ),
            )

        widths = {
            "A": 38,
            "B": 27,
            "C": 17,
            "D": 44,
            "E": 24,
            "F": 16,
            "G": 16,
            "H": 13,
            "I": 13,
            "J": 46,
            "K": 19,
            "L": 19,
            "M": 11,
            "N": 19,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        sheet.auto_filter.ref = f"A1:N{max(1, sheet.max_row)}"
        sheet.row_dimensions[1].height = 34

    def _build_history(
        self,
        sheet,
        events: Sequence[BusinessAuditEvent],
        record_by_id: dict[str, BusinessWorkflowRecord],
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.append(self.HISTORY_HEADERS)
        self._style_header(sheet["1:1"])

        for event in events:
            record = record_by_id[event.record_id]
            sheet.append(
                [
                    event.timestamp.replace(tzinfo=None),
                    event.record_id,
                    record.tender_id,
                    record.title,
                    ACTION_LABELS.get(event.action, event.action),
                    FIELD_LABELS.get(event.field, event.field or "Запись"),
                    self._display_audit_value(
                        event.field,
                        event.old_value,
                    ),
                    self._display_audit_value(
                        event.field,
                        event.new_value,
                    ),
                    event.actor,
                ]
            )
            row_number = sheet.max_row
            sheet.cell(row_number, 1).number_format = "dd.mm.yyyy hh:mm:ss"
            for cell in sheet[row_number]:
                cell.border = self._thin_border()
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        if events:
            table = Table(
                displayName="WorkflowAuditTable",
                ref=f"A1:I{sheet.max_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        widths = {
            "A": 21,
            "B": 38,
            "C": 17,
            "D": 42,
            "E": 20,
            "F": 20,
            "G": 36,
            "H": 36,
            "I": 18,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.auto_filter.ref = f"A1:I{max(1, sheet.max_row)}"
        sheet.row_dimensions[1].height = 34

    def _style_header(self, cell_range) -> None:
        """Style both a single worksheet row and a rectangular range.

        openpyxl returns ``tuple[Cell, ...]`` for ``sheet["1:1"]`` and
        ``tuple[tuple[Cell, ...], ...]`` for ranges such as ``A1:B1``.
        Normalize both shapes before applying styles.
        """
        for item in cell_range:
            cells = item if isinstance(item, (tuple, list)) else (item,)
            for cell in cells:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=self.DARK,
                )
                cell.font = Font(
                    bold=True,
                    color=self.WHITE,
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = self._thin_border()

    def _thin_border(self) -> Border:
        side = Side(style="thin", color=self.BORDER)
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )

    @staticmethod
    def _record_updated_at(
        record: BusinessWorkflowRecord,
    ) -> datetime:
        return WorkflowExcelExporter._parse_datetime(record.updated_at) or datetime.min

    @staticmethod
    def _parse_date(value: str) -> date | None:
        text = str(value or "").strip()
        if not text:
            return None
        for pattern in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_datetime(value: str) -> datetime | None:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text).replace(tzinfo=None)
        except ValueError:
            return None

    @staticmethod
    def _display_audit_value(field: str, value: str) -> str:
        if value == "":
            return "Не указано"
        if field == "status":
            return STATUS_LABELS.get(value, value)
        if field in {"total", "profit"}:
            try:
                return f"{Decimal(value):,.2f} ₽".replace(",", " ")
            except Exception:
                return value
        if field == "margin_percent":
            try:
                return f"{Decimal(value):.2f}%"
            except Exception:
                return value
        return value


__all__ = [
    "WorkflowExcelExporter",
    "WorkflowExcelExportResult",
]
