"""Validated Excel import for business workflow records."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
import re
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook

from app.repositories.business_metrics import (
    BusinessMetricsRepository,
    BusinessRecordKind,
    BusinessStatus,
    BusinessWorkflowRecord,
)


class WorkflowImportLevel(StrEnum):
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass(frozen=True, slots=True)
class WorkflowImportIssue:
    level: WorkflowImportLevel
    message: str
    field: str = ""


@dataclass(slots=True)
class WorkflowImportRow:
    source_row: int
    record_id: str = ""
    kind: BusinessRecordKind | None = None
    tender_id: str = ""
    title: str = ""
    status: BusinessStatus | None = None
    total: Decimal = Decimal("0")
    profit: Decimal = Decimal("0")
    margin_percent: Decimal = Decimal("0")
    due_date: str = ""
    file_path: str = ""
    archived: bool = False
    issues: list[WorkflowImportIssue] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not any(issue.level == WorkflowImportLevel.ERROR for issue in self.issues)

    @property
    def has_warnings(self) -> bool:
        return any(issue.level == WorkflowImportLevel.WARNING for issue in self.issues)

    @property
    def identity_key(self) -> tuple[str, str]:
        return (
            self.kind.value if self.kind is not None else "",
            self.tender_id,
        )


@dataclass(frozen=True, slots=True)
class WorkflowImportPreview:
    path: Path
    sheet_name: str
    rows: tuple[WorkflowImportRow, ...]
    fatal_issues: tuple[WorkflowImportIssue, ...] = ()

    @property
    def valid_rows(self) -> tuple[WorkflowImportRow, ...]:
        return tuple(row for row in self.rows if row.is_valid)

    @property
    def invalid_rows(self) -> tuple[WorkflowImportRow, ...]:
        return tuple(row for row in self.rows if not row.is_valid)

    @property
    def warning_rows(self) -> tuple[WorkflowImportRow, ...]:
        return tuple(row for row in self.rows if row.is_valid and row.has_warnings)

    @property
    def can_import(self) -> bool:
        return bool(self.valid_rows) and not any(
            issue.level == WorkflowImportLevel.ERROR for issue in self.fatal_issues
        )


@dataclass(frozen=True, slots=True)
class WorkflowImportResult:
    created: int = 0
    updated: int = 0
    archived: int = 0
    restored: int = 0
    skipped: int = 0
    failures: tuple[str, ...] = ()
    imported_ids: tuple[str, ...] = ()

    @property
    def imported_count(self) -> int:
        return self.created + self.updated


class WorkflowExcelImporter:
    """Read an XLSX registry, validate rows and apply valid records."""

    REQUIRED_FIELDS = {
        "kind",
        "tender_id",
        "title",
        "status",
    }

    HEADER_ALIASES = {
        "id записи": "record_id",
        "id": "record_id",
        "record id": "record_id",
        "тип": "kind",
        "тип записи": "kind",
        "kind": "kind",
        "тендер": "tender_id",
        "id тендера": "tender_id",
        "номер тендера": "tender_id",
        "tender": "tender_id",
        "наименование": "title",
        "название": "title",
        "title": "title",
        "статус": "status",
        "status": "status",
        "сумма": "total",
        "сумма руб": "total",
        "total": "total",
        "прибыль": "profit",
        "прибыль руб": "profit",
        "profit": "profit",
        "маржа": "margin_percent",
        "маржа процент": "margin_percent",
        "margin": "margin_percent",
        "срок": "due_date",
        "дата срока": "due_date",
        "due date": "due_date",
        "документ": "file_path",
        "файл": "file_path",
        "file": "file_path",
        "архив": "archived",
        "архивная запись": "archived",
        "archived": "archived",
        "дата архива": "archived_at",
    }

    KIND_LABELS = {
        "коммерческое предложение": BusinessRecordKind.PROPOSAL,
        "коммерческое предложение кп": BusinessRecordKind.PROPOSAL,
        "кп": BusinessRecordKind.PROPOSAL,
        "proposal": BusinessRecordKind.PROPOSAL,
        BusinessRecordKind.PROPOSAL.value: BusinessRecordKind.PROPOSAL,
        "смета": BusinessRecordKind.ESTIMATE,
        "estimate": BusinessRecordKind.ESTIMATE,
        BusinessRecordKind.ESTIMATE.value: BusinessRecordKind.ESTIMATE,
        "проект": BusinessRecordKind.PROJECT,
        "project": BusinessRecordKind.PROJECT,
        BusinessRecordKind.PROJECT.value: BusinessRecordKind.PROJECT,
    }

    STATUS_LABELS = {
        "черновик": BusinessStatus.DRAFT,
        "draft": BusinessStatus.DRAFT,
        "на проверке": BusinessStatus.REVIEW,
        "review": BusinessStatus.REVIEW,
        "согласовано": BusinessStatus.APPROVED,
        "approved": BusinessStatus.APPROVED,
        "готово к отправке": BusinessStatus.READY,
        "ready": BusinessStatus.READY,
        "отправлено": BusinessStatus.SENT,
        "sent": BusinessStatus.SENT,
        "принято заказчиком": BusinessStatus.ACCEPTED,
        "accepted": BusinessStatus.ACCEPTED,
        "запланировано": BusinessStatus.PLANNED,
        "planned": BusinessStatus.PLANNED,
        "в работе": BusinessStatus.ACTIVE,
        "active": BusinessStatus.ACTIVE,
        "монтаж": BusinessStatus.INSTALLATION,
        "installation": BusinessStatus.INSTALLATION,
        "пусконаладка": BusinessStatus.COMMISSIONING,
        "commissioning": BusinessStatus.COMMISSIONING,
        "завершено": BusinessStatus.COMPLETED,
        "completed": BusinessStatus.COMPLETED,
        "отменено": BusinessStatus.CANCELLED,
        "cancelled": BusinessStatus.CANCELLED,
        "canceled": BusinessStatus.CANCELLED,
        "заблокировано": BusinessStatus.BLOCKED,
        "blocked": BusinessStatus.BLOCKED,
    }

    STATUSES_BY_KIND = {
        BusinessRecordKind.ESTIMATE: {
            BusinessStatus.DRAFT,
            BusinessStatus.REVIEW,
            BusinessStatus.APPROVED,
            BusinessStatus.COMPLETED,
            BusinessStatus.BLOCKED,
            BusinessStatus.CANCELLED,
        },
        BusinessRecordKind.PROPOSAL: {
            BusinessStatus.DRAFT,
            BusinessStatus.REVIEW,
            BusinessStatus.READY,
            BusinessStatus.SENT,
            BusinessStatus.ACCEPTED,
            BusinessStatus.COMPLETED,
            BusinessStatus.BLOCKED,
            BusinessStatus.CANCELLED,
        },
        BusinessRecordKind.PROJECT: {
            BusinessStatus.PLANNED,
            BusinessStatus.ACTIVE,
            BusinessStatus.INSTALLATION,
            BusinessStatus.COMMISSIONING,
            BusinessStatus.COMPLETED,
            BusinessStatus.BLOCKED,
            BusinessStatus.CANCELLED,
        },
    }

    def preview(
        self,
        path: str | Path,
        *,
        existing_records: Sequence[BusinessWorkflowRecord] = (),
    ) -> WorkflowImportPreview:
        target = Path(path)
        fatal: list[WorkflowImportIssue] = []

        if not target.exists():
            return WorkflowImportPreview(
                path=target,
                sheet_name="",
                rows=(),
                fatal_issues=(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        f"Файл не найден: {target}",
                    ),
                ),
            )

        if target.suffix.lower() != ".xlsx":
            fatal.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    "Поддерживаются только файлы XLSX.",
                )
            )
            return WorkflowImportPreview(
                path=target,
                sheet_name="",
                rows=(),
                fatal_issues=tuple(fatal),
            )

        try:
            workbook = load_workbook(
                target,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            return WorkflowImportPreview(
                path=target,
                sheet_name="",
                rows=(),
                fatal_issues=(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        f"Не удалось открыть Excel-файл: {exc}",
                    ),
                ),
            )

        try:
            sheet = (
                workbook["Реестр"]
                if "Реестр" in workbook.sheetnames
                else workbook[workbook.sheetnames[0]]
            )
            header_row, columns = self._find_header(sheet)
            if header_row is None:
                fatal.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        "Не найдена строка заголовков реестра.",
                    )
                )
                return WorkflowImportPreview(
                    path=target,
                    sheet_name=sheet.title,
                    rows=(),
                    fatal_issues=tuple(fatal),
                )

            missing = self.REQUIRED_FIELDS - set(columns.values())
            if missing:
                labels = ", ".join(sorted(missing))
                fatal.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        f"Отсутствуют обязательные колонки: {labels}.",
                    )
                )
                return WorkflowImportPreview(
                    path=target,
                    sheet_name=sheet.title,
                    rows=(),
                    fatal_issues=tuple(fatal),
                )

            rows: list[WorkflowImportRow] = []
            for excel_row in range(header_row + 1, sheet.max_row + 1):
                values = {
                    field: sheet.cell(
                        row=excel_row,
                        column=column,
                    ).value
                    for column, field in columns.items()
                }
                if self._is_empty_row(values.values()):
                    continue
                rows.append(self._parse_row(excel_row, values))

            self._validate_duplicates(rows)
            self._validate_existing(rows, existing_records)

            if not rows:
                fatal.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.WARNING,
                        "В реестре нет строк для импорта.",
                    )
                )

            return WorkflowImportPreview(
                path=target,
                sheet_name=sheet.title,
                rows=tuple(rows),
                fatal_issues=tuple(fatal),
            )
        finally:
            workbook.close()

    def apply(
        self,
        preview: WorkflowImportPreview,
        repository: BusinessMetricsRepository,
    ) -> WorkflowImportResult:
        if not preview.can_import:
            return WorkflowImportResult(
                skipped=len(preview.rows),
                failures=("Предварительная проверка не разрешает импорт.",),
            )

        current = repository.list_records(include_archived=True)
        by_id = {record.id: record for record in current}
        by_key: dict[tuple[str, str], list[BusinessWorkflowRecord]] = {}
        for record in current:
            by_key.setdefault(
                (record.kind, record.tender_id),
                [],
            ).append(record)

        created = 0
        updated = 0
        archived_count = 0
        restored_count = 0
        skipped = len(preview.invalid_rows)
        failures: list[str] = []
        imported_ids: list[str] = []

        for row in preview.valid_rows:
            try:
                existing = self._resolve_existing(
                    row,
                    by_id=by_id,
                    by_key=by_key,
                )

                if existing is not None:
                    working = existing
                    if working.is_archived:
                        working = repository.restore_record(working.id)
                        restored_count += 1

                    working = repository.update_record(
                        working.id,
                        title=row.title,
                        total=row.total,
                        profit=row.profit,
                        margin_percent=row.margin_percent,
                        file_path=row.file_path,
                        due_date=row.due_date,
                    )
                    if row.status is not None and working.status != row.status.value:
                        working = repository.update_status(
                            working.id,
                            row.status,
                        )
                    updated += 1
                else:
                    working = repository.save_record(
                        kind=row.kind,
                        tender_id=row.tender_id,
                        title=row.title,
                        status=row.status,
                        total=row.total,
                        profit=row.profit,
                        margin_percent=row.margin_percent,
                        file_path=row.file_path,
                        due_date=row.due_date,
                    )
                    created += 1

                if row.archived and not working.is_archived:
                    working = repository.archive_record(working.id)
                    archived_count += 1

                imported_ids.append(working.id)
                by_id[working.id] = working
                by_key.setdefault(
                    (working.kind, working.tender_id),
                    [],
                )
                if all(item.id != working.id for item in by_key[(working.kind, working.tender_id)]):
                    by_key[(working.kind, working.tender_id)].append(working)
            except Exception as exc:
                skipped += 1
                failures.append(f"Строка {row.source_row}: {exc}")

        return WorkflowImportResult(
            created=created,
            updated=updated,
            archived=archived_count,
            restored=restored_count,
            skipped=skipped,
            failures=tuple(failures),
            imported_ids=tuple(imported_ids),
        )

    def _find_header(
        self,
        sheet,
    ) -> tuple[int | None, dict[int, str]]:
        best_row: int | None = None
        best_columns: dict[int, str] = {}

        for row_number in range(1, min(sheet.max_row, 15) + 1):
            columns: dict[int, str] = {}
            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(
                    row=row_number,
                    column=column,
                ).value
                key = self.HEADER_ALIASES.get(self._normalize(value))
                if key:
                    columns[column] = key

            if len(columns) > len(best_columns):
                best_row = row_number
                best_columns = columns

        if len(best_columns) < 4:
            return None, {}
        return best_row, best_columns

    def _parse_row(
        self,
        source_row: int,
        values: dict[str, Any],
    ) -> WorkflowImportRow:
        row = WorkflowImportRow(source_row=source_row)
        row.record_id = self._text(values.get("record_id"))
        row.tender_id = self._text(values.get("tender_id"))
        row.title = self._text(values.get("title"))
        row.file_path = self._text(values.get("file_path"))

        row.kind = self._parse_kind(
            values.get("kind"),
            row.issues,
        )
        row.status = self._parse_status(
            values.get("status"),
            row.issues,
        )

        row.total = self._parse_decimal(
            values.get("total"),
            field_name="Сумма",
            issues=row.issues,
            minimum=Decimal("0"),
        )
        row.profit = self._parse_decimal(
            values.get("profit"),
            field_name="Прибыль",
            issues=row.issues,
            minimum=Decimal("0"),
        )
        row.margin_percent = self._parse_decimal(
            values.get("margin_percent"),
            field_name="Маржа",
            issues=row.issues,
            minimum=Decimal("-100"),
            maximum=Decimal("1000"),
        )
        row.due_date = self._parse_date(
            values.get("due_date"),
            row.issues,
        )
        row.archived = self._parse_bool(
            values.get("archived"),
            row.issues,
        )

        if not row.tender_id:
            row.issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    "Не указан тендер.",
                    "tender_id",
                )
            )
        if not row.title:
            row.issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    "Не указано наименование.",
                    "title",
                )
            )

        if (
            row.kind is not None
            and row.status is not None
            and row.status not in self.STATUSES_BY_KIND[row.kind]
        ):
            row.issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    (f"Статус «{row.status.value}» не подходит для типа «{row.kind.value}»."),
                    "status",
                )
            )

        if row.profit > row.total and row.total > 0:
            row.issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.WARNING,
                    "Прибыль превышает общую сумму.",
                    "profit",
                )
            )

        if row.total > 0 and row.margin_percent == 0 and row.profit > 0:
            row.margin_percent = (row.profit / row.total * Decimal("100")).quantize(Decimal("0.01"))
            row.issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.INFO,
                    "Маржа рассчитана автоматически.",
                    "margin_percent",
                )
            )

        return row

    def _validate_duplicates(
        self,
        rows: Sequence[WorkflowImportRow],
    ) -> None:
        ids: dict[str, list[WorkflowImportRow]] = {}
        keys: dict[tuple[str, str], list[WorkflowImportRow]] = {}

        for row in rows:
            if row.record_id:
                ids.setdefault(row.record_id, []).append(row)
            if row.kind is not None and row.tender_id:
                keys.setdefault(row.identity_key, []).append(row)

        for record_id, duplicates in ids.items():
            if len(duplicates) <= 1:
                continue
            for row in duplicates:
                row.issues.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        f"ID «{record_id}» повторяется в файле.",
                        "record_id",
                    )
                )

        for key, duplicates in keys.items():
            if len(duplicates) <= 1:
                continue
            for row in duplicates:
                row.issues.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        (f"Повторяется сочетание типа и тендера: {key[0]} / {key[1]}."),
                    )
                )

    def _validate_existing(
        self,
        rows: Sequence[WorkflowImportRow],
        existing_records: Sequence[BusinessWorkflowRecord],
    ) -> None:
        by_id = {record.id: record for record in existing_records}
        by_key: dict[tuple[str, str], list[BusinessWorkflowRecord]] = {}
        for record in existing_records:
            by_key.setdefault(
                (record.kind, record.tender_id),
                [],
            ).append(record)

        for row in rows:
            if row.record_id and row.record_id in by_id:
                existing = by_id[row.record_id]
                if row.kind is not None and (
                    existing.kind != row.kind.value or existing.tender_id != row.tender_id
                ):
                    row.issues.append(
                        WorkflowImportIssue(
                            WorkflowImportLevel.ERROR,
                            ("ID найден, но тип или тендер не совпадает с существующей записью."),
                            "record_id",
                        )
                    )
                else:
                    row.issues.append(
                        WorkflowImportIssue(
                            WorkflowImportLevel.INFO,
                            "Существующая запись будет обновлена.",
                        )
                    )
                continue

            matches = by_key.get(row.identity_key, [])
            if len(matches) > 1:
                row.issues.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.ERROR,
                        (
                            "В базе несколько записей с таким типом "
                            "и тендером. Укажите точный ID записи."
                        ),
                    )
                )
            elif len(matches) == 1:
                row.issues.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.INFO,
                        "Запись будет обновлена по типу и тендеру.",
                    )
                )
            elif row.record_id:
                row.issues.append(
                    WorkflowImportIssue(
                        WorkflowImportLevel.WARNING,
                        ("ID не найден в базе; будет создана новая запись с новым внутренним ID."),
                        "record_id",
                    )
                )

    @staticmethod
    def _resolve_existing(
        row: WorkflowImportRow,
        *,
        by_id: dict[str, BusinessWorkflowRecord],
        by_key: dict[
            tuple[str, str],
            list[BusinessWorkflowRecord],
        ],
    ) -> BusinessWorkflowRecord | None:
        if row.record_id and row.record_id in by_id:
            return by_id[row.record_id]

        matches = by_key.get(row.identity_key, [])
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise ValueError("Найдено несколько совпадающих записей.")
        return None

    def _parse_kind(
        self,
        value: Any,
        issues: list[WorkflowImportIssue],
    ) -> BusinessRecordKind | None:
        normalized = self._normalize(value)
        kind = self.KIND_LABELS.get(normalized)
        if kind is None:
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"Неизвестный тип записи: {value!s}.",
                    "kind",
                )
            )
        return kind

    def _parse_status(
        self,
        value: Any,
        issues: list[WorkflowImportIssue],
    ) -> BusinessStatus | None:
        normalized = self._normalize(value)
        status = self.STATUS_LABELS.get(normalized)
        if status is None:
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"Неизвестный статус: {value!s}.",
                    "status",
                )
            )
        return status

    def _parse_decimal(
        self,
        value: Any,
        *,
        field_name: str,
        issues: list[WorkflowImportIssue],
        minimum: Decimal | None = None,
        maximum: Decimal | None = None,
    ) -> Decimal:
        if value in {None, ""}:
            return Decimal("0")

        if isinstance(value, bool):
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"{field_name}: ожидается число.",
                )
            )
            return Decimal("0")

        text = str(value).strip()
        text = (
            text.replace("\u00a0", "")
            .replace(" ", "")
            .replace("₽", "")
            .replace("%", "")
            .replace(",", ".")
        )
        try:
            result = Decimal(text)
        except (InvalidOperation, ValueError):
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"{field_name}: неверное число «{value}».",
                )
            )
            return Decimal("0")

        if minimum is not None and result < minimum:
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"{field_name}: значение меньше {minimum}.",
                )
            )
        if maximum is not None and result > maximum:
            issues.append(
                WorkflowImportIssue(
                    WorkflowImportLevel.ERROR,
                    f"{field_name}: значение больше {maximum}.",
                )
            )
        return result

    def _parse_date(
        self,
        value: Any,
        issues: list[WorkflowImportIssue],
    ) -> str:
        if value in {None, ""}:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        text = str(value).strip()
        for pattern in (
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(text, pattern).date().isoformat()
            except ValueError:
                continue

        issues.append(
            WorkflowImportIssue(
                WorkflowImportLevel.ERROR,
                f"Неверная дата срока: «{value}».",
                "due_date",
            )
        )
        return ""

    def _parse_bool(
        self,
        value: Any,
        issues: list[WorkflowImportIssue],
    ) -> bool:
        if value in {None, ""}:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)

        normalized = self._normalize(value)
        if normalized in {"да", "yes", "true", "1", "архив"}:
            return True
        if normalized in {"нет", "no", "false", "0", "активная"}:
            return False

        issues.append(
            WorkflowImportIssue(
                WorkflowImportLevel.ERROR,
                f"Неверное значение архива: «{value}».",
                "archived",
            )
        )
        return False

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _normalize(value: Any) -> str:
        text = str(value or "").strip().lower().replace("ё", "е")
        text = re.sub(r"[₽%(),.:;_\-/]+", " ", text)
        return " ".join(text.split())

    @staticmethod
    def _is_empty_row(values: Iterable[Any]) -> bool:
        return not any(value is not None and str(value).strip() for value in values)


__all__ = [
    "WorkflowExcelImporter",
    "WorkflowImportIssue",
    "WorkflowImportLevel",
    "WorkflowImportPreview",
    "WorkflowImportResult",
    "WorkflowImportRow",
]
